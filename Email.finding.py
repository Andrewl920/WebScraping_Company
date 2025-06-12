from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl import load_workbook
import time
import re
import Excel as excel

class EmailFinding:
    def __init__(self, filename = r"C:\Users\Administrator\Desktop\WebScraping_Company\Air Con and Electrical (Yellow Page).xlsx"):
        service = Service(r"C:\Users\Administrator\Desktop\chromedriver-win64\chromedriver.exe")
        self.driver = webdriver.Chrome(service=service)
        self.driver.maximize_window()
        self.wait = WebDriverWait(self.driver, 10)
        self.base_url = "https://www.google.com/maps"
        self.filename = filename
        self.workbook = load_workbook(self.filename)
        self.sheet = self.workbook.active
        
    def search_email(self, url):
        if url == "NA":
            email = "NA"
            return email
        
        # url = "http://" + url

        try: 
            self.driver.set_page_load_timeout(5)
            self.driver.get(url)
            time.sleep(5)
        except TimeoutException: 
            email = "NA"
            return email
        except WebDriverException:
            email = "NA"
            return email
        
        try:
            emails = self.find_email(url)
            if len(emails) == 0:
                contact_element = self.driver.find_element(By.XPATH, "//*[contains(text(), 'Contact')]")
                contact_element.click()
                time.sleep(3)
                emails = self.find_email(url)
                if len(emails) == 0:
                    url = url + "/contact"
                    self.driver.get(url)
                    time.sleep(5)
                    emails = self.find_email(url)
                
                if len(emails) == 0:
                    emails = "NA"
                
        except:
            emails = "NA"
        
        return emails


    def find_url(self, number=2):
        return str(self.sheet.cell(row = number, column = 5).value)
    
    #find email from the link
    def find_email_from_href(self):
        # self.driver.get("http://rileybuilding.com.au/contact")
        mailto_elements = self.driver.find_elements(By.XPATH, '//a[starts-with(@href, "mailto:")]')
        emails = set()
        for elem in mailto_elements:
            href = elem.get_attribute('href')
            if href and "mailto:" in href:
                email = href.split("mailto:")[1]
                emails.add(email)
        
        return list(emails)
    
    def find_email(self, url):
        body = self.driver.find_element(By.TAG_NAME, "body").text
        emails = re.findall(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', body)
        emails = list(set(emails))
        if len(emails) == 0:
            emails = self.find_email_from_href()

        print(len(emails))
        return emails
    
    def append_email(self, email, number):
        if type(email) == list:
            col_add_number = 0
            for i in range(len(email)):
                self.sheet.cell(row = number, column = 5 + col_add_number).value = str(email[i])
                col_add_number += 1
            
        else:
            self.sheet.cell(row = number, column = 5).value = str(email)
        
        self.workbook.save(self.filename)
    
    def close_driver(self):
        self.driver.quit()
        
if __name__ == "__main__":
    
    excel_file = excel.Excel()
    last_row = excel_file.find_last_row("QLD")
    total_number = int(last_row[1:])
    
    for company in range(2, total_number):
        email_search = EmailFinding()
        url = excel_file.workbook["QLD"].cell(row = company, column = 4).value
        email = email_search.search_email(url)
        time.sleep(5)
        email_search.append_email(email, company)
        print("finish" + str(company - 1))

        email_search.driver.close()


    # email_search = EmailFinding()
    
    # email = email_search.search_email("http://www.tannerselectrical.com/")
    
            


        