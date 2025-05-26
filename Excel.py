from openpyxl import load_workbook 

class Excel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(filename = self.file_path)

    #find the first cell that not filled with value
    def find_last_row(self, sheetname):
        for row in self.workbook[sheetname].iter_rows(max_col=1):
            for cell in row:
                if cell.value is not None:
                    last_cell = cell.coordinate
                    col = last_cell[:1]
                    row = int(last_cell[1:]) + 1
                    
                    
        return(col + str(row))

    #fill in the information of the company according to the state
    def fill_in_value(self, sheetname, coordinate, company_info):
        row = coordinate[1:] #get the row number

        self.workbook[sheetname]["A"+str(row)] = company_info["Name"]
        self.workbook[sheetname]["B"+str(row)] = company_info["Phone"]
        self.workbook[sheetname]["C"+str(row)] = company_info["Address"]
        self.workbook[sheetname]["D"+str(row)] = company_info["Website"]

        self.workbook.save(self.file_path)

    #find out the state name
    def state_name(self, company_info):
        # print(company_info["Address"])
        state = ["QLD", "NSW", "VIC", "SA", "ACT", "WA", "NT", "TAS"]
        if company_info["Address"].split(" ")[1] in state:
            state = company_info["Address"].split(" ")[1]
        elif company_info["Address"].split(" ")[2] in state:
            state = company_info["Address"].split(" ")[2]
        elif company_info["Address"].split(" ")[3] in state:
            state = company_info["Address"].split(" ")[3]
        elif company_info["Address"].split(" ")[4] in state:
            state = company_info["Address"].split(" ")[4]
        else:
            return "No state"
        
        return state

    #check if there is duplicate
    def check_duplicate(self, sheetname, company_info):
        company_name = company_info["Name"]
        company_address = company_info["Address"]
        company_phone = company_info["Phone"]
        company_website = company_info["Website"]
        for row in self.workbook[sheetname].iter_rows(min_row=1, values_only=True):
            cell_a = row[0] # Column A
            cell_b = row[1] # Column B
            cell_c = row[2] # Column C
            cell_d = row[3] # Column D
            
            if company_name == cell_a and company_phone == cell_b and company_address == cell_c and company_website == cell_d:
                return True
            else:
                continue
        
        return False
